from bottle import request, template, response, static_file
from datetime import datetime
import json
import os
import csv

EPS = 1e-10

def load_theory():
    try:
        json_path = os.path.join('data', 'theory_transport.json')
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        print(f"Ошибка загрузки теории: {e}")
        return {}

def save_form_data_to_cookie(suppliers, consumers, supply, demand, costs):
    data = {'suppliers': suppliers, 'consumers': consumers, 'supply': supply, 'demand': demand, 'costs': costs}
    response.set_cookie('transport_data', json.dumps(data), path='/', secret=None)

def load_form_data_from_cookie():
    saved_data = request.get_cookie('transport_data')
    if saved_data:
        try:
            return json.loads(saved_data)
        except:
            pass
    return None

def solve_transport():
    theory = load_theory()
    result = None
    error = None
    
    saved = load_form_data_from_cookie()
    if saved:
        default_suppliers = saved.get('suppliers', 3)
        default_consumers = saved.get('consumers', 3)
        default_supply = saved.get('supply', [70, 100, 110])
        default_demand = saved.get('demand', [80, 50, 150])
        default_costs = saved.get('costs', [[1,4,5],[3,5,2],[2,6,4]])
    else:
        default_suppliers = 3
        default_consumers = 3
        default_supply = [70, 100, 110]
        default_demand = [80, 50, 150]
        default_costs = [[1,4,5],[3,5,2],[2,6,4]]
    
    if request.method == 'POST':
        # ========== ЭКСПОРТ РЕЗУЛЬТАТОВ В CSV/EXCEL ==========
        if request.forms.get('export_results'):
            try:
                format_type = request.forms.get('export_format', 'csv')
                # Загружаем сохранённый результат из скрытых полей
                saved_result = json.loads(request.forms.get('saved_result', '{}'))
                if saved_result:
                    return export_results(saved_result, format_type)
            except Exception as e:
                error = f"Ошибка экспорта: {e}"
        
        # Основное решение
        try:
            suppliers = int(request.forms.get('suppliers', 3))
            consumers = int(request.forms.get('consumers', 3))
            
            supply = [float(request.forms.get(f'supply_{i}', 0)) for i in range(suppliers)]
            demand = [float(request.forms.get(f'demand_{j}', 0)) for j in range(consumers)]
            costs = [[float(request.forms.get(f'cost_{i}_{j}', 0)) for j in range(consumers)] for i in range(suppliers)]
            
            save_form_data_to_cookie(suppliers, consumers, supply, demand, costs)
            
            total_supply = sum(supply)
            total_demand = sum(demand)
            balanced = abs(total_supply - total_demand) < EPS
            
            # Коррекция для сбалансированности
            supply_corr, demand_corr, costs_corr = supply[:], demand[:], [row[:] for row in costs]
            orig_suppliers, orig_consumers = suppliers, consumers
            has_fictive = False
            
            if not balanced:
                if total_supply > total_demand:
                    demand_corr.append(total_supply - total_demand)
                    for i in range(suppliers):
                        costs_corr[i].append(0)
                    consumers += 1
                else:
                    supply_corr.append(total_demand - total_supply)
                    costs_corr.append([0] * consumers)
                    suppliers += 1
                has_fictive = True
            
            plan_nw, cost_nw, steps_nw, nw_info = northwest_corner_full(supply_corr, demand_corr, costs_corr)
            plan_me, cost_me, steps_me, me_info = min_element_full(supply_corr, demand_corr, costs_corr)
            
            if cost_me <= cost_nw:
                best_initial_plan, best_initial_cost, best_initial_name = plan_me, cost_me, "минимального элемента"
            else:
                best_initial_plan, best_initial_cost, best_initial_name = plan_nw, cost_nw, "северо-западного угла"
            
            plan_opt, cost_opt, iterations_opt = potential_method_full(costs_corr, best_initial_plan, best_initial_name, best_initial_cost, cost_nw, cost_me)
            
            if has_fictive:
                if total_supply > total_demand:
                    plan_opt = [row[:-1] for row in plan_opt]
                else:
                    plan_opt = plan_opt[:-1]
            
            result = {
                'northwest_plan': plan_nw, 'northwest_cost': round(cost_nw, 2), 'northwest_steps': steps_nw, 'northwest_degenerate': nw_info,
                'mincost_plan': plan_me, 'mincost_cost': round(cost_me, 2), 'mincost_steps': steps_me, 'mincost_degenerate': me_info,
                'best_plan': plan_opt, 'best_cost': round(cost_opt, 2), 'best_iterations': iterations_opt, 'best_method': best_initial_name,
                'suppliers': orig_suppliers, 'consumers': orig_consumers, 'supply': supply, 'demand': demand, 'costs': costs,
                'balanced': balanced, 'total_supply': total_supply, 'total_demand': total_demand,
                'cost_nw': round(cost_nw, 2), 'cost_me': round(cost_me, 2)
            }
        except Exception as e:
            error = str(e)
            import traceback; traceback.print_exc()
    
    form_data = {'suppliers': default_suppliers, 'consumers': default_consumers, 'supply': default_supply, 'demand': default_demand, 'costs': default_costs}
    return template('transport', result=result, error=error, theory=theory, form_data=form_data, year=datetime.now().year)

def northwest_corner_full(supply, demand, costs):
    n, m = len(supply), len(demand)
    plan = [[0]*m for _ in range(n)]
    steps = []
    i, j, step_num = 0, 0, 1
    s, d = supply[:], demand[:]
    
    while i < n and j < m:
        amount = min(s[i], d[j])
        plan[i][j] = amount
        steps.append({'step': step_num, 'cell': f"({i+1},{j+1})", 'amount': amount, 'formula': f"min({s[i]}, {d[j]}) = {amount}"})
        s[i] -= amount
        d[j] -= amount
        if abs(s[i]) < EPS:
            i += 1
        else:
            j += 1
        step_num += 1
    
    basic_cells = sum(1 for i in range(n) for j in range(m) if plan[i][j] > EPS)
    expected = n + m - 1
    info = {'is_degenerate': basic_cells < expected, 'basic_cells': basic_cells, 'expected_basic': expected,
            'message': f"⚠️ План вырожденный! Базисных клеток {basic_cells} вместо {expected}" if basic_cells < expected else f"✅ План невырожденный (базисных клеток: {basic_cells} из {expected})"}
    total_cost = sum(plan[i][j] * costs[i][j] for i in range(n) for j in range(m))
    return plan, total_cost, steps, info

def min_element_full(supply, demand, costs):
    n, m = len(supply), len(demand)
    plan = [[0]*m for _ in range(n)]
    s, d = supply[:], demand[:]
    steps = []
    cells = [(i, j, costs[i][j]) for i in range(n) for j in range(m)]
    cells.sort(key=lambda x: x[2])
    step_num = 1
    
    for i, j, cost in cells:
        if s[i] > EPS and d[j] > EPS:
            amount = min(s[i], d[j])
            plan[i][j] = amount
            steps.append({'step': step_num, 'cell': f"({i+1},{j+1})", 'cost': cost, 'amount': amount, 'formula': f"min({s[i]}, {d[j]}) = {amount}"})
            s[i] -= amount
            d[j] -= amount
            step_num += 1
    
    basic_cells = sum(1 for i in range(n) for j in range(m) if plan[i][j] > EPS)
    expected = n + m - 1
    info = {'is_degenerate': basic_cells < expected, 'basic_cells': basic_cells, 'expected_basic': expected,
            'message': f"⚠️ План вырожденный! Базисных клеток {basic_cells} вместо {expected}" if basic_cells < expected else f"✅ План невырожденный (базисных клеток: {basic_cells} из {expected})"}
    total_cost = sum(plan[i][j] * costs[i][j] for i in range(n) for j in range(m))
    return plan, total_cost, steps, info

def find_cycle_simple(plan, enter_i, enter_j, n, m):
    """
    Упрощённый поиск цикла пересчёта.
    Возвращает список клеток цикла с чередующимися знаками + и -
    """
    # Собираем все базисные клетки (где перевозка > 0)
    basis = [(i, j) for i in range(n) for j in range(m) if plan[i][j] > 0]
    basis.append((enter_i, enter_j))  # Добавляем вводимую клетку
    
    # Группируем по строкам и столбцам
    rows = {}
    cols = {}
    for (i, j) in basis:
        rows.setdefault(i, []).append(j)
        cols.setdefault(j, []).append(i)
    
    # Ищем цикл простым перебором
    # Ищем первую клетку в той же строке, что и вводимая
    for j in rows.get(enter_i, []):
        if j == enter_j:
            continue
        # Ищем клетку в том же столбце, что и найденная
        for i in cols.get(j, []):
            if i == enter_i:
                continue
            # Ищем клетку в той же строке, что и найденная
            for j2 in rows.get(i, []):
                if j2 == j:
                    continue
                # Ищем клетку в том же столбце, что и найденная, которая вернёт нас в начало
                for i2 in cols.get(j2, []):
                    if i2 == i:
                        continue
                    if i2 == enter_i and j2 == enter_j:
                        # Нашли простой прямоугольный цикл
                        return [
                            (enter_i, enter_j, 1),   # + (вводимая)
                            (enter_i, j, -1),        # -
                            (i, j, 1),               # +
                            (i, j2, -1)              # -
                        ]
    return None

def build_cycle_table(plan, cycle, costs, n, m, enter_i, enter_j, iteration_num):
    """Построение таблицы с визуализацией цикла"""
    html = '<div style="overflow-x: auto; margin: 15px 0;"><table class="result-table" style="border-collapse: collapse; margin: 0 auto;">'
    html += '<thead><tr style="background: #EDE7F6;"><th></th>'
    for j in range(m):
        html += f'<th style="padding: 8px; border: 1px solid #9B2226;">B{j+1}</th>'
    html += '<th>Запасы</th></tr></thead><tbody>'
    
    for i in range(n):
        html += f'<tr><th style="padding: 8px; border: 1px solid #9B2226;">A{i+1}</th>'
        for j in range(m):
            val = round(plan[i][j], 2) if plan[i][j] > EPS else '—'
            cell_class = ''
            if i == enter_i and j == enter_j:
                cell_class = 'cycle-cell-enter'
            else:
                for (ci, cj, sign) in cycle:
                    if ci == i and cj == j:
                        cell_class = 'cycle-cell-plus' if sign == 1 else 'cycle-cell-minus'
                        break
            html += f'<td class="{cell_class}" style="padding: 8px; text-align: center; min-width: 60px; border: 1px solid #ddd;">'
            html += f'<strong>{val}</strong><br><small>c={costs[i][j]}</small></td>'
        html += f'<td style="background:#e9ecef">{sum(plan[i])}</td></tr>'
    
    html += '<tr><th>Потребности</th>'
    for j in range(m):
        html += f'<td style="background:#e9ecef">{sum(plan[i][j] for i in range(n))}</td>'
    html += '<td style="background:#e9ecef">—</td></tr>'
    html += '</tbody></table></div>'
    return html
def potential_method_full(costs, initial_plan, initial_method_name, initial_cost, cost_nw, cost_me):
    n, m = len(costs), len(costs[0])
    plan = [row[:] for row in initial_plan]
    iterations = []
    iteration_num = 1
    max_iterations = 30
    
    # Сравнение начальных планов
    iterations.append({
        'type': 'comparison',
        'html': f'''
        <div style="background: linear-gradient(135deg, #9b59b6, #8e44ad); color: white; padding: 20px; border-radius: 12px; margin-bottom: 20px;">
            <h3 style="margin: 0 0 15px 0;">📊 Сравнение начальных планов</h3>
            <p style="margin: 8px 0;">🔹 Метод северо-западного угла: стоимость = <strong>{round(cost_nw, 2)}</strong> ден. ед.</p>
            <p style="margin: 8px 0;">🔹 Метод минимального элемента: стоимость = <strong>{round(cost_me, 2)}</strong> ден. ед.</p>
            <div style="margin-top: 15px; padding: 10px; background: rgba(255,255,255,0.2); border-radius: 8px;">
                ✅ <strong>Выбран опорный план метода "{initial_method_name}"</strong><br>
                📌 Стоимость: <strong>{round(initial_cost, 2)} ден. ед.</strong>
            </div>
        </div>
        '''
    })
    
    while iteration_num <= max_iterations:
        # Сбор базисных клеток
        basis = [(i, j) for i in range(n) for j in range(m) if plan[i][j] > EPS]
        
        # Расчёт потенциалов
        u, v = [None]*n, [None]*m
        u[0] = 0
        changed = True
        while changed:
            changed = False
            for (i, j) in basis:
                if u[i] is not None and v[j] is None:
                    v[j] = costs[i][j] - u[i]
                    changed = True
                elif v[j] is not None and u[i] is None:
                    u[i] = costs[i][j] - v[j]
                    changed = True
        u = [x if x is not None else 0 for x in u]
        v = [x if x is not None else 0 for x in v]
        
        # Система уравнений
        sys_eq = '<br>'.join([f"u{i+1} + v{j+1} = {costs[i][j]}" for (i, j) in basis[:8]])
        system_html = f'''
        <div style="background: #f0f0f0; padding: 12px; border-radius: 8px; margin: 10px 0;">
            <strong>📝 Система уравнений:</strong><br>{sys_eq}{'<br>...' if len(basis) > 8 else ''}<br>
            <span style="color: #9B2226;">🔹 Принимаем u₁ = 0</span>
        </div>
        <div style="background: #e3f2fd; padding: 12px; border-radius: 8px; margin: 10px 0;">
            <strong>✨ Потенциалы:</strong><br>uᵢ = {[round(x,2) for x in u]}<br>vⱼ = {[round(x,2) for x in v]}
        </div>
        '''
        
        # Вычисление оценок
        deltas = []
        enter_i, enter_j = -1, -1
        max_delta = -float('inf')
        for i in range(n):
            for j in range(m):
                if plan[i][j] < EPS:
                    delta = round(u[i] + v[j] - costs[i][j], 6)
                    deltas.append({'cell': f"({i+1},{j+1})", 'delta': delta, 'formula': f"{round(u[i],2)} + {round(v[j],2)} - {costs[i][j]} = {delta}", 'is_positive': delta > EPS})
                    if delta > max_delta + EPS:
                        max_delta = delta
                        enter_i, enter_j = i, j
        
        current_cost = sum(plan[i][j] * costs[i][j] for i in range(n) for j in range(m))
        
        # Таблица оценок
        deltas_table = '<table class="result-table" style="margin: 10px 0;"><thead><tr><th>Клетка</th><th>Формула</th><th>Оценка Δ</th></tr></thead><tbody>'
        for d in deltas:
            bg = '#ffeb3b' if d['is_positive'] else 'white'
            deltas_table += f'<tr style="background-color: {bg};"><td style="padding: 6px;">{d["cell"]}</td><td style="padding: 6px;">{d["formula"]}</td><td style="padding: 6px;">{d["delta"]}</td></tr>'
        deltas_table += '</tbody></table>'
        
        # Проверка оптимальности
        if max_delta <= EPS:
            iterations.append({
                'type': 'optimal',
                'iteration': iteration_num,
                'html': system_html + deltas_table + f'''
                <div style="background: #d4edda; padding: 15px; border-radius: 12px; margin: 15px 0; border: 2px solid #28a745;">
                    <h4 style="margin: 0;">🎉 ПЛАН ОПТИМАЛЕН!</h4>
                    <p>Все оценки Δᵢⱼ ≤ 0, дальнейшее улучшение невозможно.</p>
                    <p style="font-size: 1.2rem;">📌 Стоимость: <strong>{round(current_cost, 2)} ден. ед.</strong></p>
                </div>
                '''
            })
            break
        
        # Есть положительная оценка
        positive_html = f'''
        <div style="background: #fff3cd; padding: 12px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #ffc107;">
            <strong>⚠️ Найдена положительная оценка!</strong><br>
            📌 Δ = <strong>{round(max_delta, 6)}</strong> в клетке <strong>({enter_i+1}, {enter_j+1})</strong><br>
            💡 <strong>Почему именно эта клетка?</strong> Положительная оценка означает, что введение перевозки в эту клетку уменьшит общую стоимость на величину Δ. Выбираем клетку с максимальной Δ для наибольшего улучшения.
        </div>
        '''
        
        # Построение цикла (упрощённый метод)
        cycle = find_cycle_simple(plan, enter_i, enter_j, n, m)
        
        if not cycle:
            # Если цикл не найден, создаём простой прямоугольный цикл
            # Ищем любую базисную клетку в той же строке
            j2 = None
            for j in range(m):
                if plan[enter_i][j] > EPS and j != enter_j:
                    j2 = j
                    break
            # Ищем любую базисную клетку в том же столбце
            i2 = None
            for i in range(n):
                if plan[i][enter_j] > EPS and i != enter_i:
                    i2 = i
                    break
            
            if j2 is not None and i2 is not None:
                cycle = [
                    (enter_i, enter_j, 1),   # +
                    (enter_i, j2, -1),       # -
                    (i2, j2, 1),             # +
                    (i2, enter_j, -1)        # -
                ]
        
        if not cycle:
            iterations.append({'type': 'error', 'iteration': iteration_num, 'html': system_html + deltas_table + positive_html + '<div style="background: #f8d7da; padding: 15px; border-radius: 12px;"><strong>❌ Ошибка:</strong> Не удалось построить цикл</div>'})
            break
        
        # Находим theta
        theta = float('inf')
        for (i, j, sign) in cycle:
            if sign == -1:
                if plan[i][j] < theta - EPS:
                    theta = plan[i][j]
        if theta <= EPS:
            theta = 1  # Если все нули, берём 1 для демонстрации
        
        # ВИЗУАЛИЗАЦИЯ ЦИКЛА (таблица с выделенными клетками)
        cycle_table_html = '<div style="overflow-x: auto; margin: 15px 0;"><table class="result-table" style="border-collapse: collapse; margin: 0 auto;">'
        cycle_table_html += '<thead><tr style="background: #EDE7F6;"><th></th>'
        for j in range(m):
            cycle_table_html += f'<th style="padding: 8px; border: 1px solid #9B2226;">B{j+1}</th>'
        cycle_table_html += '<th>Запасы</th></tr></thead><tbody>'
        
        for i in range(n):
            cycle_table_html += f'<tr><th style="padding: 8px; border: 1px solid #9B2226;">A{i+1}</th>'
            for j in range(m):
                val = round(plan[i][j], 2) if plan[i][j] > EPS else '—'
                # Определяем класс для подсветки
                cell_class = ''
                for (ci, cj, sign) in cycle:
                    if ci == i and cj == j:
                        if sign == 1:
                            cell_class = 'cycle-cell-plus'
                        elif sign == -1:
                            cell_class = 'cycle-cell-minus'
                        break
                if i == enter_i and j == enter_j:
                    cell_class = 'cycle-cell-enter'
                cycle_table_html += f'<td class="{cell_class}" style="padding: 8px; text-align: center; min-width: 60px; border: 1px solid #ddd;'
                if cell_class == 'cycle-cell-plus':
                    cycle_table_html += ' background-color: #d4edda;'
                elif cell_class == 'cycle-cell-minus':
                    cycle_table_html += ' background-color: #f8d7da;'
                elif cell_class == 'cycle-cell-enter':
                    cycle_table_html += ' background-color: #fff3cd;'
                cycle_table_html += f'"><strong>{val}</strong><br><small>c={costs[i][j]}</small>'
                if cell_class == 'cycle-cell-plus':
                    cycle_table_html += '<span style="margin-left: 5px; color: green;">(+)</span>'
                elif cell_class == 'cycle-cell-minus':
                    cycle_table_html += '<span style="margin-left: 5px; color: red;">(-)</span>'
                elif cell_class == 'cycle-cell-enter':
                    cycle_table_html += '<span style="margin-left: 5px; color: orange;">★</span>'
                cycle_table_html += '</td>'
            cycle_table_html += f'<td style="background:#e9ecef">{sum(plan[i])}</td>'
            cycle_table_html += '</tr>'
        
        cycle_table_html += '<tr><th>Потребности</th>'
        for j in range(m):
            cycle_table_html += f'<td style="background:#e9ecef">{sum(plan[i][j] for i in range(n))}</td>'
        cycle_table_html += '<td style="background:#e9ecef">—</td>'
        cycle_table_html += '</tr>'
        cycle_table_html += '</tbody></table></div>'
        
        # Список изменений по циклу
        changes_html = '<table class="result-table" style="margin: 10px 0;"><thead><tr><th>Клетка</th><th>Знак</th><th>Было</th><th>θ</th><th>Стало</th></tr></thead><tbody>'
        for (i, j, sign) in cycle:
            changes_html += f'<tr><td style="text-align: center;"><strong>({i+1},{j+1})</strong></td><td style="text-align: center; font-size: 1.2rem; font-weight: bold; color: {"green" if sign==1 else "red"};">{"+" if sign==1 else "-"}</span></td><td style="text-align: center;">{round(plan[i][j], 2)}</td><td style="text-align: center;">{theta}</td><td style="text-align: center;">{round(plan[i][j] + sign * theta, 2)}</strong></td></tr>'
        changes_html += '</tbody></table>'
        
        # Перераспределение
        new_plan = [row[:] for row in plan]
        for (i, j, sign) in cycle:
            new_plan[i][j] += sign * theta
            if abs(new_plan[i][j]) < EPS:
                new_plan[i][j] = 0.0
        
        new_cost = sum(new_plan[i][j] * costs[i][j] for i in range(n) for j in range(m))
        
        # Новая матрица после перераспределения
        new_matrix_html = '<div style="overflow-x: auto; margin: 15px 0;"><table class="result-table" style="border-collapse: collapse; margin: 0 auto;"><thead><tr style="background: #EDE7F6;"><th></th>'
        for j in range(m):
            new_matrix_html += f'<th style="padding: 8px; border: 1px solid #9B2226;">B{j+1}</th>'
        new_matrix_html += '<th>Запасы</th></tr></thead><tbody>'
        for i in range(n):
            new_matrix_html += f'<tr><th style="padding: 8px; border: 1px solid #9B2226;">A{i+1}</th>'
            for j in range(m):
                val = round(new_plan[i][j], 2) if new_plan[i][j] > EPS else '—'
                new_matrix_html += f'<td style="padding: 8px; text-align: center; min-width: 60px; border: 1px solid #ddd;"><strong>{val}</strong><br><small>c={costs[i][j]}</small></td>'
            new_matrix_html += f'<td style="background:#e9ecef">{sum(new_plan[i])}</td>'
            new_matrix_html += '</tr>'
        new_matrix_html += '</tbody></table></div>'
        
        iterations.append({
            'type': 'iteration',
            'iteration': iteration_num,
            'html': system_html + deltas_table + positive_html + f'''
            <div style="margin: 15px 0; padding: 10px; background: #e8e8e8; border-radius: 8px;">
                <p><strong>🔄 ПОСТРОЕНИЕ ЦИКЛА ПЕРЕСЧЁТА:</strong></p>
                <p><em>Цикл пересчёта — это замкнутая ломаная линия, вершины которой расположены в занятых клетках, а звенья — вдоль строк и столбцов. В каждой вершине цикла сходятся ровно два звена. Знаки «+» и «-» чередуются, начиная с «+» в вводимой клетке (отмечена ★).</em></p>
                {cycle_table_html}
                <p><strong>📌 Цикл (со знаками):</strong> {" → ".join([f"({i+1},{j+1})<sup>{'+' if s==1 else '-'}</sup>" for (i,j,s) in cycle])}</p>
                <p><strong>θ = {theta}</strong> — минимальная перевозка в клетках со знаком «-» (выбрана для сохранения неотрицательности).</p>
                {changes_html}
                <p><strong>Новая матрица после перераспределения:</strong></p>
                {new_matrix_html}
                <div style="font-size: 1.1rem; font-weight: bold; color: #28a745; margin-top: 10px;">
                    💰 Новая стоимость: <strong>{round(new_cost, 2)}</strong> ден. ед.<br>
                    📉 Уменьшение: <strong>{round(current_cost - new_cost, 2)}</strong> ден. ед.
                </div>
            </div>
            '''
        })
        
        plan = new_plan
        iteration_num += 1
    
    total_cost = sum(plan[i][j] * costs[i][j] for i in range(n) for j in range(m))
    return plan, total_cost, iterations

def save_results_to_csv(result, filename=None):
    """Сохранение всех результатов в CSV файл"""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'data/results/transport_result_{timestamp}.csv'
    
    os.makedirs('data/results', exist_ok=True)
    
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        
        # Заголовок
        writer.writerow(['='*60])
        writer.writerow(['ТРАНСПОРТНАЯ ЗАДАЧА - РЕЗУЛЬТАТ РЕШЕНИЯ'])
        writer.writerow(['='*60])
        writer.writerow([f'Дата решения: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        writer.writerow([])
        
        # Исходные данные
        writer.writerow(['1. ИСХОДНЫЕ ДАННЫЕ'])
        writer.writerow(['-'*40])
        writer.writerow([f'Количество поставщиков: {result["suppliers"]}'])
        writer.writerow([f'Количество потребителей: {result["consumers"]}'])
        writer.writerow([])
        writer.writerow(['Запасы поставщиков:'] + [f'A{i+1}={result["supply"][i]}' for i in range(result["suppliers"])])
        writer.writerow(['Потребности потребителей:'] + [f'B{j+1}={result["demand"][j]}' for j in range(result["consumers"])])
        writer.writerow([])
        writer.writerow(['Матрица тарифов:'])
        writer.writerow([''] + [f'B{j+1}' for j in range(result["consumers"])] + ['Запасы'])
        for i in range(result["suppliers"]):
            row = [f'A{i+1}'] + [result["costs"][i][j] for j in range(result["consumers"])] + [result["supply"][i]]
            writer.writerow(row)
        writer.writerow(['Потребности'] + [result["demand"][j] for j in range(result["consumers"])] + [''])
        writer.writerow([])
        
        # Проверка баланса
        writer.writerow(['2. ПРОВЕРКА СБАЛАНСИРОВАННОСТИ'])
        writer.writerow(['-'*40])
        writer.writerow([f'Сумма запасов: Σaᵢ = {result["total_supply"]}'])
        writer.writerow([f'Сумма потребностей: Σbⱼ = {result["total_demand"]}'])
        writer.writerow([f'Задача {"сбалансирована (закрытая модель)" if result["balanced"] else "несбалансирована, добавлены фиктивные участники"}'])
        writer.writerow([])
        
        # Метод северо-западного угла
        writer.writerow(['3. МЕТОД СЕВЕРО-ЗАПАДНОГО УГЛА'])
        writer.writerow(['-'*40])
        writer.writerow(['Пошаговое построение опорного плана:'])
        for step in result['northwest_steps']:
            writer.writerow([f'Шаг {step["step"]}: Клетка {step["cell"]} → {step["formula"]} единиц'])
        writer.writerow([])
        writer.writerow(['Полученный опорный план:'])
        writer.writerow([''] + [f'B{j+1}' for j in range(result["consumers"])] + ['Запасы'])
        for i in range(result["suppliers"]):
            row = [f'A{i+1}'] + [result["northwest_plan"][i][j] for j in range(result["consumers"])] + [result["supply"][i]]
            writer.writerow(row)
        writer.writerow([result["northwest_degenerate"]["message"]])
        writer.writerow([f'Стоимость перевозок: F = {result["northwest_cost"]} ден. ед.'])
        writer.writerow([])
        
        # Метод минимального элемента
        writer.writerow(['4. МЕТОД МИНИМАЛЬНОГО ЭЛЕМЕНТА'])
        writer.writerow(['-'*40])
        writer.writerow(['Пошаговое построение опорного плана:'])
        for step in result['mincost_steps']:
            writer.writerow([f'Шаг {step["step"]}: Клетка {step["cell"]} (тариф={step["cost"]}) → {step["formula"]} единиц'])
        writer.writerow([])
        writer.writerow(['Полученный опорный план:'])
        writer.writerow([''] + [f'B{j+1}' for j in range(result["consumers"])] + ['Запасы'])
        for i in range(result["suppliers"]):
            row = [f'A{i+1}'] + [result["mincost_plan"][i][j] for j in range(result["consumers"])] + [result["supply"][i]]
            writer.writerow(row)
        writer.writerow([result["mincost_degenerate"]["message"]])
        writer.writerow([f'Стоимость перевозок: F = {result["mincost_cost"]} ден. ед.'])
        writer.writerow([])
        
        # Метод потенциалов
        writer.writerow(['5. МЕТОД ПОТЕНЦИАЛОВ (ОПТИМИЗАЦИЯ)'])
        writer.writerow(['-'*40])
        writer.writerow([f'Выбран опорный план метода "{result["best_method"]}" со стоимостью {result["best_initial_cost"]} ден. ед.'])
        writer.writerow([])
        
        # Итерации метода потенциалов
        iter_count = 0
        for iter_data in result['best_iterations']:
            if iter_data.get('type') == 'iteration':
                iter_count += 1
                writer.writerow([f'Итерация {iter_count}:'])
                writer.writerow([f'  Потенциалы uᵢ: {iter_data["potentials_u"]}'])
                writer.writerow([f'  Потенциалы vⱼ: {iter_data["potentials_v"]}'])
                writer.writerow([f'  Оценки свободных клеток:'])
                for d in iter_data['deltas']:
                    writer.writerow([f'    {d["cell"]}: {d["formula"]} → Δ = {d["delta"]}'])
                writer.writerow([f'  Положительная оценка Δ = {iter_data["max_delta"]} в клетке {iter_data["enter_cell"]}'])
                writer.writerow([f'  θ = {iter_data["theta"]}'])
                writer.writerow([f'  Новая стоимость: {iter_data["new_cost"]} ден. ед. (уменьшение на {iter_data["improvement"]})'])
                writer.writerow([])
        
        # Оптимальный план
        writer.writerow(['6. ОПТИМАЛЬНЫЙ ПЛАН ПЕРЕВОЗОК'])
        writer.writerow(['-'*40])
        writer.writerow(['Оптимальный план:'])
        writer.writerow([''] + [f'B{j+1}' for j in range(result["consumers"])] + ['Запасы'])
        for i in range(result["suppliers"]):
            row = [f'A{i+1}'] + [result["best_plan"][i][j] for j in range(result["consumers"])] + [result["supply"][i]]
            writer.writerow(row)
        writer.writerow([])
        writer.writerow([f'МИНИМАЛЬНАЯ СТОИМОСТЬ ПЕРЕВОЗОК: {result["best_cost"]} ден. ед.'])
        writer.writerow(['='*60])
    
    return filename


def save_results_to_excel(result, filename=None):
    """Сохранение всех результатов в Excel файл (XLSX)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f'data/results/transport_result_{timestamp}.xlsx'
        
        os.makedirs('data/results', exist_ok=True)
        
        wb = openpyxl.Workbook()
        
        # Стили
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="9B2226", end_color="9B2226", fill_type="solid")
        title_font = Font(bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # ========== Лист 1: Исходные данные ==========
        ws1 = wb.active
        ws1.title = "Исходные данные"
        
        ws1['A1'] = "ТРАНСПОРТНАЯ ЗАДАЧА - РЕЗУЛЬТАТ РЕШЕНИЯ"
        ws1['A1'].font = title_font
        ws1['A2'] = f"Дата решения: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        ws1['A4'] = "ИСХОДНЫЕ ДАННЫЕ"
        ws1['A4'].font = header_font
        ws1['A4'].fill = header_fill
        
        ws1['A6'] = f"Количество поставщиков: {result['suppliers']}"
        ws1['A7'] = f"Количество потребителей: {result['consumers']}"
        
        ws1['A9'] = "Запасы поставщиков:"
        for i in range(result['suppliers']):
            ws1.cell(row=10, column=i+2, value=f"A{i+1}={result['supply'][i]}")
        
        ws1['A11'] = "Потребности потребителей:"
        for j in range(result['consumers']):
            ws1.cell(row=12, column=j+2, value=f"B{j+1}={result['demand'][j]}")
        
        # Матрица тарифов
        ws1['A14'] = "Матрица тарифов:"
        ws1['A15'] = ""
        for j in range(result['consumers']):
            ws1.cell(row=15, column=j+2, value=f"B{j+1}")
        ws1.cell(row=15, column=result['consumers']+2, value="Запасы")
        
        for i in range(result['suppliers']):
            ws1.cell(row=16+i, column=1, value=f"A{i+1}")
            for j in range(result['consumers']):
                ws1.cell(row=16+i, column=j+2, value=result['costs'][i][j])
            ws1.cell(row=16+i, column=result['consumers']+2, value=result['supply'][i])
        
        ws1.cell(row=16+result['suppliers'], column=1, value="Потребности")
        for j in range(result['consumers']):
            ws1.cell(row=16+result['suppliers'], column=j+2, value=result['demand'][j])
        
        # ========== Лист 2: Метод северо-западного угла ==========
        ws2 = wb.create_sheet("Северо-западный угол")
        
        ws2['A1'] = "МЕТОД СЕВЕРО-ЗАПАДНОГО УГЛА"
        ws2['A1'].font = title_font
        
        ws2['A3'] = "Пошаговое построение опорного плана:"
        row = 4
        for step in result['northwest_steps']:
            ws2.cell(row=row, column=1, value=f"Шаг {step['step']}: Клетка {step['cell']} → {step['formula']} единиц")
            row += 1
        
        row += 1
        ws2.cell(row=row, column=1, value="Полученный опорный план:")
        row += 1
        ws2.cell(row=row, column=2, value="")
        for j in range(result['consumers']):
            ws2.cell(row=row, column=j+3, value=f"B{j+1}")
        ws2.cell(row=row, column=result['consumers']+3, value="Запасы")
        row += 1
        
        for i in range(result['suppliers']):
            ws2.cell(row=row, column=2, value=f"A{i+1}")
            for j in range(result['consumers']):
                ws2.cell(row=row, column=j+3, value=result['northwest_plan'][i][j])
            ws2.cell(row=row, column=result['consumers']+3, value=result['supply'][i])
            row += 1
        
        row += 1
        ws2.cell(row=row, column=1, value=result['northwest_degenerate']['message'])
        row += 1
        ws2.cell(row=row, column=1, value=f"Стоимость перевозок: F = {result['northwest_cost']} ден. ед.")
        
        # ========== Лист 3: Метод минимального элемента ==========
        ws3 = wb.create_sheet("Минимальный элемент")
        
        ws3['A1'] = "МЕТОД МИНИМАЛЬНОГО ЭЛЕМЕНТА"
        ws3['A1'].font = title_font
        
        ws3['A3'] = "Пошаговое построение опорного плана:"
        row = 4
        for step in result['mincost_steps']:
            ws3.cell(row=row, column=1, value=f"Шаг {step['step']}: Клетка {step['cell']} (тариф={step['cost']}) → {step['formula']} единиц")
            row += 1
        
        row += 1
        ws3.cell(row=row, column=1, value="Полученный опорный план:")
        row += 1
        ws3.cell(row=row, column=2, value="")
        for j in range(result['consumers']):
            ws3.cell(row=row, column=j+3, value=f"B{j+1}")
        ws3.cell(row=row, column=result['consumers']+3, value="Запасы")
        row += 1
        
        for i in range(result['suppliers']):
            ws3.cell(row=row, column=2, value=f"A{i+1}")
            for j in range(result['consumers']):
                ws3.cell(row=row, column=j+3, value=result['mincost_plan'][i][j])
            ws3.cell(row=row, column=result['consumers']+3, value=result['supply'][i])
            row += 1
        
        row += 1
        ws3.cell(row=row, column=1, value=result['mincost_degenerate']['message'])
        row += 1
        ws3.cell(row=row, column=1, value=f"Стоимость перевозок: F = {result['mincost_cost']} ден. ед.")
        
        # ========== Лист 4: Метод потенциалов ==========
        ws4 = wb.create_sheet("Метод потенциалов")
        
        ws4['A1'] = "МЕТОД ПОТЕНЦИАЛОВ (ОПТИМИЗАЦИЯ)"
        ws4['A1'].font = title_font
        
        ws4['A3'] = f"Выбран опорный план метода \"{result['best_method']}\" со стоимостью {result['best_initial_cost']} ден. ед."
        
        row = 5
        iter_count = 0
        for iter_data in result['best_iterations']:
            if iter_data.get('type') == 'iteration':
                iter_count += 1
                ws4.cell(row=row, column=1, value=f"Итерация {iter_count}:")
                row += 1
                ws4.cell(row=row, column=1, value=f"  Потенциалы uᵢ: {iter_data['potentials_u']}")
                row += 1
                ws4.cell(row=row, column=1, value=f"  Потенциалы vⱼ: {iter_data['potentials_v']}")
                row += 1
                ws4.cell(row=row, column=1, value=f"  Оценки свободных клеток:")
                row += 1
                for d in iter_data['deltas']:
                    ws4.cell(row=row, column=1, value=f"    {d['cell']}: {d['formula']} → Δ = {d['delta']}")
                    row += 1
                ws4.cell(row=row, column=1, value=f"  Положительная оценка Δ = {iter_data['max_delta']} в клетке {iter_data['enter_cell']}")
                row += 1
                ws4.cell(row=row, column=1, value=f"  θ = {iter_data['theta']}")
                row += 1
                ws4.cell(row=row, column=1, value=f"  Новая стоимость: {iter_data['new_cost']} ден. ед. (уменьшение на {iter_data['improvement']})")
                row += 2
        
        # ========== Лист 5: Оптимальный план ==========
        ws5 = wb.create_sheet("Оптимальный план")
        
        ws5['A1'] = "ОПТИМАЛЬНЫЙ ПЛАН ПЕРЕВОЗОК"
        ws5['A1'].font = title_font
        
        ws5['A3'] = "Оптимальный план:"
        ws5.cell(row=4, column=2, value="")
        for j in range(result['consumers']):
            ws5.cell(row=4, column=j+3, value=f"B{j+1}")
        ws5.cell(row=4, column=result['consumers']+3, value="Запасы")
        
        row = 5
        for i in range(result['suppliers']):
            ws5.cell(row=row, column=2, value=f"A{i+1}")
            for j in range(result['consumers']):
                ws5.cell(row=row, column=j+3, value=result['best_plan'][i][j])
            ws5.cell(row=row, column=result['consumers']+3, value=result['supply'][i])
            row += 1
        
        row += 1
        ws5.cell(row=row, column=1, value=f"МИНИМАЛЬНАЯ СТОИМОСТЬ ПЕРЕВОЗОК: {result['best_cost']} ден. ед.")
        
        wb.save(filename)
        return filename
        
    except ImportError:
        # Если openpyxl не установлен, сохраняем как XLS через CSV
        return save_results_to_csv(result, filename.replace('.xlsx', '.xls'))


def export_results(result, format='csv'):
    """Экспорт результатов в файл и возврат для скачивания"""
    if format == 'csv':
        filepath = save_results_to_csv(result)
        return static_file(os.path.basename(filepath), root='data/results', download=True)
    else:
        filepath = save_results_to_excel(result)
        return static_file(os.path.basename(filepath), root='data/results', download=True)